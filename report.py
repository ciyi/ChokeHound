"""
Reporting module for ChokeHound -BloodHound CE Choke Points Analyzer

This module handles Excel file creation, formatting, and report generation.
"""

import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# Import configuration
import config
import risk_config


# Mapping of relationship types to their BloodHound documentation URL paths
RELATIONSHIP_TYPE_URL_MAP = {
    "AbuseTGTDelegation": "abuse-tgt-delegation",
    "ADCSESC1": "adcs-esc1",
    "ADCSESC10a": "adcs-esc10a",
    "ADCSESC10b": "adcs-esc10b",
    "ADCSESC13": "adcs-esc13",
    "ADCSESC3": "adcs-esc3",
    "ADCSESC4": "adcs-esc4",
    "ADCSESC6a": "adcs-esc6a",
    "ADCSESC6b": "adcs-esc6b",
    "ADCSESC9a": "adcs-esc9a",
    "ADCSESC9b": "adcs-esc9b",
    "AddAllowedToAct": "add-allowed-to-act",
    "AddKeyCredentialLink": "add-key-credential-link",
    "AddMember": "add-member",
    "AddSelf": "add-self",
    "AdminTo": "admin-to",
    "AllExtendedRights": "all-extended-rights",
    "AllowedToAct": "allowed-to-act",
    "AllowedToDelegate": "allowed-to-delegate",
    "CanPSRemote": "can-psremote",
    "CanRDP": "can-rdp",
    "CoerceAndRelayNTLMToADCS": "coerce-and-relay-ntlm-to-adcs",
    "CoerceAndRelayNTLMToLDAP": "coerce-and-relay-ntlm-to-ldap",
    "CoerceAndRelayNTLMToLDAPS": "coerce-and-relay-ntlm-to-ldaps",
    "CoerceAndRelayNTLMToSMB": "coerce-and-relay-ntlm-to-smb",
    "CoerceToTGT": "coerce-to-tgt",
    "DCFor": "dc-for",
    "DCSync": "dcsync",
    "DumpSMSAPassword": "dump-smsa-password",
    "ExecuteDCOM": "execute-dcom",
    "ForceChangePassword": "force-change-password",
    "GenericAll": "generic-all",
    "GenericWrite": "generic-write",
    "GoldenCert": "golden-cert",
    "HasSIDHistory": "has-sid-history",
    "HasSession": "has-session",
    "HasTrustKeys": "has-trust-keys",
    "MemberOf": "member-of",
    "Owns": "owns",
    "OwnsLimitedRights": "owns-limited-rights",
    "ReadGMSAPassword": "read-gmsa-password",
    "ReadLAPSPassword": "read-laps-password",
    "SameForestTrust": "same-forest-trust",
    "SpoofSIDHistory": "spoof-sid-history",
    "SQLAdmin": "sql-admin",
    "SyncedToEntraUser": "synced-to-entra-user",
    "SyncLAPSPassword": "sync-laps-password",
    "WriteAccountRestrictions": "write-account-restrictions",
    "WriteDacl": "write-dacl",
    "WriteGPLink": "write-gplink",
    "WriteOwner": "write-owner",
    "WriteOwnerLimitedRights": "write-owner-limited-rights",
    "WriteSPN": "write-spn",
}


def convert_relationship_type_to_url(relationship_type):
    """
    Convert relationship type name to BloodHound documentation URL format.
    
    Uses a predefined mapping dictionary for accurate URL generation.
    Falls back to lowercase conversion if relationship type is not in the map.
    
    Args:
        relationship_type: Relationship type string
        
    Returns:
        URL-friendly string (lowercase with hyphens)
    """
    relationship_type = str(relationship_type).strip()
    
    # Check if we have an exact mapping
    if relationship_type in RELATIONSHIP_TYPE_URL_MAP:
        return RELATIONSHIP_TYPE_URL_MAP[relationship_type]
    
    # Fallback: convert to lowercase (for any unmapped relationship types)
    # This handles edge cases where new relationship types might be added
    return relationship_type.lower()


def add_relationship_type_hyperlinks(worksheet, df):
    """
    Add hyperlinks to RelationshipType column cells pointing to BloodHound documentation.
    
    Args:
        worksheet: openpyxl Worksheet object
        df: pandas DataFrame that was written to the worksheet
    """
    if df.empty or 'RelationshipType' not in df.columns:
        return
    
    # Find the RelationshipType column index
    relationship_type_col = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == 'RelationshipType':
            relationship_type_col = idx
            break
    
    if relationship_type_col is None:
        return
    
    # Get the column letter
    col_letter = get_column_letter(relationship_type_col)
    
    # Base URL for BloodHound edge documentation
    base_url = "https://bloodhound.specterops.io/resources/edges/"
    
    # Add hyperlinks to each data row (skip header row 1)
    for row_idx, relationship_type in enumerate(df['RelationshipType'], start=2):
        if pd.notna(relationship_type) and relationship_type:
            # Convert relationship type to URL format
            relationship_type_str = str(relationship_type).strip()
            url_suffix = convert_relationship_type_to_url(relationship_type_str)
            url = base_url + url_suffix
            
            # Get the cell
            cell = worksheet[f"{col_letter}{row_idx}"]
            
            # Set hyperlink
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
            cell.value = relationship_type_str  # Keep the original value


def color_risk_column(worksheet, df):
    """
    Color the RiskScore column based on risk levels:
    - Yellow (low risk): 1 to 30
    - Orange (medium risk): 31 to 59
    - Red (high risk): 60 to 100
    
    Args:
        worksheet: openpyxl Worksheet object
        df: pandas DataFrame that was written to the worksheet
    """
    if df.empty or 'RiskScore' not in df.columns:
        return
    
    # Find the RiskScore column index
    risk_score_col = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == 'RiskScore':
            risk_score_col = idx
            break
    
    if risk_score_col is None:
        return
    
    # Get the column letter
    col_letter = get_column_letter(risk_score_col)
    
    # Define fill colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for high risk
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for medium risk
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for low risk
    
    # Color each data row (skip header row 1)
    for row_idx, risk_score in enumerate(df['RiskScore'], start=2):
        if pd.notna(risk_score):
            try:
                risk_value = float(risk_score)
                cell = worksheet[f"{col_letter}{row_idx}"]
                
                # Apply color based on risk level
                if risk_value >= 60:
                    cell.fill = red_fill
                elif risk_value >= 31:
                    cell.fill = orange_fill
                else:
                    cell.fill = yellow_fill
            except (ValueError, TypeError):
                # Skip if risk_score is not a valid number
                continue


def format_sheet_as_table(worksheet, df):
    """
    Format a worksheet as an Excel table with filterable/sortable headers.
    
    Args:
        worksheet: openpyxl Worksheet object
        df: pandas DataFrame that was written to the worksheet
    """
    if df.empty:
        return
    
    # Get the data range
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    if max_row < 2:  # Need at least header + 1 data row
        return
    
    start_cell = "A1"
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    table_range = f"{start_cell}:{end_cell}"
    
    # Create table name (Excel table names must be unique and valid)
    table_name = f"Table_{worksheet.title[:20].replace(' ', '_')}"
    # Remove invalid characters
    table_name = ''.join(c for c in table_name if c.isalnum() or c == '_')
    
    # Create table
    table = Table(displayName=table_name, ref=table_range)
    
    # Add default table style
    style = TableStyleInfo(
        name="TableStyleLight8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Add table to worksheet
    worksheet.add_table(table)
    
    # Format header row: set text color to white
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_documentation_sheet(workbook, dataframes):
    """
    Create a documentation sheet explaining what each sheet contains and column meanings.
    
    Args:
        workbook: openpyxl Workbook object
        dataframes: Dictionary mapping sheet names to pandas DataFrames
    """
    # Create documentation sheet (insert at position 1, after Cover)
    doc_sheet = workbook.create_sheet("Documentation", 1)
    
    # Title
    doc_sheet['A1'] = "Report Documentation"
    doc_sheet['A1'].font = Font(size=16, bold=True)
    doc_sheet['A1'].alignment = Alignment(horizontal='left')
    doc_sheet.merge_cells('A1:D1')
    
    row = 3
    
    # Introduction
    doc_sheet[f'A{row}'] = "This document explains the contents of each sheet and the meaning of each column."
    doc_sheet[f'A{row}'].font = Font(italic=True)
    doc_sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
    doc_sheet.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # Define column descriptions for each sheet
    sheet_column_docs = {
        "Critical Choke Points Risk": {
            "description": "Prioritized critical choke points that connect non-Tier-0 objects with Tier-0 objects. Results are ranked by risk score (highest risk first) and number of affected attack paths.",
            "columns": {
                "ID": "Unique numeric identifier for each choke point (1 = highest risk, 2 = second highest, etc.)",
                "SourceName": "Name of the source object (non-Tier-0) that has a relationship to a Tier-0 target",
                "SourceType": "Type of the source object (e.g., User, Group, Computer, Container, OU, GPO)",
                "SourceObjectID": "Security Identifier (SID) of the source object. Used internally for identifying common default groups.",
                "SourceDN": "Distinguished Name (DN) of the source object in Active Directory",
                "RelationshipType": "Type of relationship/privilege from source to target (e.g., Owns, GenericAll, AdminTo). Click hyperlinks for detailed explanations.",
                "TargetName": "Name of the target Tier-0 object",
                "TargetType": "Type of the target object (e.g., User, Group, Computer, Domain)",
                "TargetObjectID": "Security Identifier (SID) of the target object",
                "TargetDN": "Distinguished Name (DN) of the target object in Active Directory",
                "AffectedAttackPaths": "Number of unique non-Tier-0 origins that can reach this choke point (up to the configured hop limit). Higher numbers indicate more attack paths are blocked by securing this choke point.",
                "RiskScore": "Normalized risk score on a scale of 1-100, where 100 represents the highest possible risk. Calculated based on source object type, relationship type, target object type, and number of affected attack paths. Raw risk scores are documented in the risk calculation log file (generated with --log-risk)."
            }
        },
        "Direct relationships into Tier0": {
            "description": "Ranking of relationship types (edges) that directly connect to Tier-0 objects. This identifies which types of privileges are most commonly used to access Tier-0, helping identify edge-type choke points.",
            "columns": {
                "RelationshipType": "Type of relationship/privilege (e.g., Owns, GenericAll, AdminTo)",
                "DistinctSourceTargetPairs": "Number of unique source-target pairs connected by this relationship type",
                "TotalEdges": "Total number of edges of this relationship type connecting to Tier-0 objects"
            }
        },
        "Source nodes into T0 Rank": {
            "description": "Top non-Tier-0 nodes ranked by the number of distinct Tier-0 targets they directly connect to. These are immediate choke points - objects that have direct relationships to many Tier-0 objects.",
            "columns": {
                "SourceNode": "Name of the source object (non-Tier-0)",
                "SourceType": "Type of the source object (e.g., User, Group, Computer)",
                "SourceObjectID": "Security Identifier (SID) of the source object",
                "SourceDN": "Distinguished Name (DN) of the source object in Active Directory",
                "DistinctTier0Targets": "Number of distinct Tier-0 objects this source directly connects to",
                "RelationshipTypes": "List of relationship types used by this source to connect to Tier-0 objects"
            }
        }
    }
    
    # Generate documentation for each sheet
    for sheet_name in dataframes.keys():
        # Skip if sheet not in documentation
        if sheet_name not in sheet_column_docs:
            continue
            
        sheet_doc = sheet_column_docs[sheet_name]
        
        # Sheet name header
        doc_sheet[f'A{row}'] = f"Sheet: {sheet_name}"
        doc_sheet[f'A{row}'].font = Font(size=14, bold=True)
        doc_sheet[f'A{row}'].alignment = Alignment(horizontal='left')
        row += 1
        
        # Sheet description
        doc_sheet[f'A{row}'] = "Description:"
        doc_sheet[f'A{row}'].font = Font(bold=True)
        doc_sheet[f'B{row}'] = sheet_doc["description"]
        doc_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        doc_sheet.merge_cells(f'B{row}:D{row}')
        row += 2
        
        # Column descriptions header
        doc_sheet[f'A{row}'] = "Column Descriptions:"
        doc_sheet[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        # Column headers
        doc_sheet[f'A{row}'] = "Column Name"
        doc_sheet[f'A{row}'].font = Font(bold=True)
        doc_sheet[f'B{row}'] = "Description"
        doc_sheet[f'B{row}'].font = Font(bold=True)
        doc_sheet.merge_cells(f'B{row}:D{row}')
        row += 1
        
        # Get actual columns from dataframe (handle case where sheet might not exist)
        actual_columns = []
        if sheet_name in dataframes:
            actual_columns = list(dataframes[sheet_name].columns)
        
        # List columns that exist in the documentation
        for col_name, col_desc in sheet_doc["columns"].items():
            # Only show columns that actually exist in the dataframe (or are always present like ID)
            if col_name in actual_columns or col_name == "ID":
                doc_sheet[f'A{row}'] = col_name
                doc_sheet[f'A{row}'].font = Font(bold=True)
                doc_sheet[f'B{row}'] = col_desc
                doc_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                doc_sheet.merge_cells(f'B{row}:D{row}')
                row += 1
        
        # Add spacing between sheets
        row += 2
    
    # Adjust column widths
    doc_sheet.column_dimensions['A'].width = 25
    doc_sheet.column_dimensions['B'].width = 60
    doc_sheet.column_dimensions['C'].width = 10
    doc_sheet.column_dimensions['D'].width = 10
    
    # Add some spacing
    for i in range(1, row + 2):
        doc_sheet.row_dimensions[i].height = 20


def create_cover_sheet(workbook, output_filename, domains=None):
    """
    Create a cover sheet for the Excel report.
    
    Args:
        workbook: openpyxl Workbook object
        output_filename: Name of the output file
        domains: Optional list of domain names to display
    """
    # Create cover sheet
    cover_sheet = workbook.create_sheet("Cover", 0)  # Insert at the beginning
    
    # Hide gridlines in the cover sheet
    cover_sheet.sheet_view.showGridLines = False
    
    # Try to add logo if it exists
    logo_paths = [
        os.path.join(os.path.dirname(__file__), "logo.png"),
        os.path.join(os.path.dirname(__file__), "logo.jpg"),
        os.path.join(os.path.dirname(__file__), "logo.jpeg"),
        os.path.join(os.path.dirname(__file__), "ChokeHound_logo.png"),
        os.path.join(os.path.dirname(__file__), "ChokeHound_logo.jpg"),
    ]
    
    logo_added = False
    logo_height = 120  # Height in pixels
    
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                # Scale logo to appropriate size (maintain aspect ratio)
                aspect_ratio = img.width / img.height
                img.height = logo_height
                img.width = int(logo_height * aspect_ratio)
                
                # Center logo above the header
                # Header spans columns A-D, so we center the logo within that range
                # Using column B as anchor point (roughly centered between A-D)
                cover_sheet.add_image(img, 'B1')
                logo_added = True
                # Adjust row heights to accommodate logo and spacing
                cover_sheet.row_dimensions[1].height = logo_height * 0.75  # Logo row
                cover_sheet.row_dimensions[2].height = 10  # Small spacing
                cover_sheet.row_dimensions[3].height = 10  # Small spacing
                cover_sheet.row_dimensions[4].height = 10  # Small spacing
                break
            except Exception as e:
                print(f"[WARNING] Could not add logo from {logo_path}: {e}")
                continue
    
    # Title row - header will be at row 6 if logo was added, row 2 if not
    # (moved everything starting from row 5 down by one row)
    title_row = 6 if logo_added else 2
    
    # Title
    cover_sheet[f'A{title_row}'] = "ChokeHound - BloodHound CE Choke Points Analysis Report"
    cover_sheet[f'A{title_row}'].font = Font(size=16, bold=True)
    cover_sheet[f'A{title_row}'].alignment = Alignment(horizontal='center')
    cover_sheet.merge_cells(f'A{title_row}:D{title_row}')
    
    # Report metadata
    row = title_row + 2
    cover_sheet[f'A{row}'] = "Report Information"
    cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
    
    row += 1
    cover_sheet[f'A{row}'] = "Generated:"
    cover_sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Domain information
    if domains:
        row += 1
        cover_sheet[f'A{row}'] = "Active Directory Domains:"
        cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        domains_text = ", ".join(domains)
        cover_sheet[f'A{row}'] = domains_text
        cover_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        cover_sheet.merge_cells(f'A{row}:D{row}')
    
    # Description section
    row += 2
    cover_sheet[f'A{row}'] = "Report Description"
    cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
    
    row += 1
    description = (
        "This report identifies choke points in Active Directory - critical privilege edges "
        "that connect non-Tier-0 objects to Tier-0 objects. Choke points represent optimal "
        "locations to block the largest number of attack paths."
    )
    cover_sheet[f'A{row}'] = description
    cover_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    cover_sheet.merge_cells(f'A{row}:D{row + 2}')
    
    # Adjust column widths
    cover_sheet.column_dimensions['A'].width = 35
    cover_sheet.column_dimensions['B'].width = 50
    cover_sheet.column_dimensions['C'].width = 15
    cover_sheet.column_dimensions['D'].width = 15
    
    # Add some spacing
    for i in range(1, row + 2):
        cover_sheet.row_dimensions[i].height = 20


def generate_risk_log(risk_breakdowns, output_filename, log_filename):
    """
    Generate a detailed log file explaining risk calculations for each choke point.
    
    Args:
        risk_breakdowns: List of dictionaries containing risk calculation breakdowns
        output_filename: Name of the Excel output file
        log_filename: Name of the log file to generate
    """
    try:
        with open(log_filename, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Risk Calculation Log - ChokeHound - BloodHound CE Choke Points Analysis\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Excel Report: {output_filename}\n")
            f.write(f"Total Choke Points: {len(risk_breakdowns)}\n\n")
            
            # Calculate and display min/max risk scores
            min_score, max_score = risk_config.calculate_risk_score_range()
            f.write("Risk Score Range:\n")
            f.write("-" * 80 + "\n")
            f.write(f"  Theoretical Minimum Risk Score: {min_score:.2f}\n")
            f.write(f"  Theoretical Maximum Risk Score: {max_score:.2f}\n")
            f.write(f"  Normalized Range: 1-100 (where {min_score:.2f} = 1 and {max_score:.2f} = 100)\n\n")
            
            f.write("Risk Calculation Formula:\n")
            f.write("-" * 80 + "\n")
            f.write("Raw Risk Score = (SourceObjectWeight × SourceObjectCategory) +\n")
            f.write("                 (RelationshipTypeWeight × RelationshipTypeCategory) +\n")
            f.write("                 (TargetObjectWeight × TargetObjectCategory) +\n")
            f.write("                 (AffectedAttackPathsWeight × PathsMultiplier × 10)\n\n")
            
            f.write("Normalization Formula:\n")
            f.write("-" * 80 + "\n")
            f.write("Normalized Risk Score (1-100) = 1 + ((RawRiskScore - MinScore) / (MaxScore - MinScore)) × 99\n")
            f.write(f"Where MinScore = {min_score:.2f} and MaxScore = {max_score:.2f}\n\n")
            
            f.write("Component Weights:\n")
            f.write("-" * 80 + "\n")
            weights = risk_config.RISK_WEIGHTS
            f.write(f"  - Source Object Weight: {weights['source_object']}\n")
            f.write(f"  - Relationship Type Weight: {weights['relationship_type']}\n")
            f.write(f"  - Target Object Weight: {weights['target_object']}\n")
            f.write(f"  - Affected Attack Paths Weight: {weights['affected_attack_paths']}\n\n")
            
            # Display risk category ranges
            f.write("Risk Category Value Ranges:\n")
            f.write("-" * 80 + "\n")
            f.write("  Source Object Categories: ")
            source_vals = risk_config.SOURCE_OBJECT_CATEGORIES.values()
            f.write(f"{min(source_vals)} to {max(source_vals)} (1-10 scale)\n")
            f.write("  Relationship Type Categories: ")
            rel_vals = risk_config.RELATIONSHIP_TYPE_CATEGORIES.values()
            f.write(f"{min(rel_vals)} to {max(rel_vals)} (1-10 scale)\n")
            f.write("  Target Object Categories: ")
            target_vals = risk_config.TARGET_OBJECT_CATEGORIES.values()
            f.write(f"{min(target_vals)} to {max(target_vals)} (1-10 scale)\n")
            f.write("  Affected Attack Paths Multipliers: ")
            path_mults = [m for _, _, m in risk_config.AFFECTED_ATTACK_PATHS_MULTIPLIERS]
            f.write(f"{min(path_mults):.1f} to {max(path_mults):.1f}\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("DETAILED RISK CALCULATIONS\n")
            f.write("=" * 80 + "\n\n")
            
            for idx, bd in enumerate(risk_breakdowns, 1):
                f.write(f"Choke Point #{idx}\n")
                f.write("-" * 80 + "\n")
                f.write(f"Source: {bd['source_name']} ({bd['source_type']})\n")
                f.write(f"  Risk Category: {bd['source_risk_category']}\n")
                f.write(f"  Weight: {bd['source_weight']}\n")
                f.write(f"  Component: {bd['source_weight']} × {bd['source_risk_category']} = {bd['source_component']}\n\n")
                
                f.write(f"Relationship: {bd['relationship_type']}\n")
                f.write(f"  Risk Category: {bd['relationship_risk_category']}\n")
                f.write(f"  Weight: {bd['relationship_weight']}\n")
                f.write(f"  Component: {bd['relationship_weight']} × {bd['relationship_risk_category']} = {bd['relationship_component']}\n\n")
                
                f.write(f"Target: {bd['target_name']} ({bd['target_type']})\n")
                f.write(f"  Risk Category: {bd['target_risk_category']}\n")
                f.write(f"  Weight: {bd['target_weight']}\n")
                f.write(f"  Component: {bd['target_weight']} × {bd['target_risk_category']} = {bd['target_component']}\n\n")
                
                f.write(f"Affected Attack Paths: {bd['affected_paths']}\n")
                f.write(f"  Multiplier: {bd['paths_multiplier']}\n")
                f.write(f"  Weight: {bd['paths_weight']}\n")
                f.write(f"  Component: {bd['paths_weight']} × {bd['paths_multiplier']} × 10 = {bd['paths_component']}\n\n")
                
                f.write(f"TOTAL RISK SCORE: {bd['source_component']} + {bd['relationship_component']} + ")
                f.write(f"{bd['target_component']} + {bd['paths_component']} = {bd['total_risk_score']}\n")
                
                # Add normalized risk score (1-100 scale)
                normalized_score = risk_config.normalize_risk_score(bd['total_risk_score'])
                f.write(f"NORMALIZED RISK SCORE (1-100): {normalized_score}\n\n")
                f.write("=" * 80 + "\n\n")
        
        print(f"[OK] Risk calculation log generated: {log_filename}")
    except Exception as e:
        print(f"[WARNING] Error generating risk log: {e}")


def create_excel_report(dataframes, output_filename, risk_breakdowns=None, enable_logging=False, domains=None):
    """
    Create an Excel report from dataframes with formatting and cover sheet.
    
    Args:
        dataframes: Dictionary mapping sheet names to pandas DataFrames
        output_filename: Name of the output Excel file
        risk_breakdowns: Optional list of risk breakdown dictionaries for logging
        enable_logging: If True, generate risk calculation log file
        domains: Optional list of domain names to display on cover sheet
        
    Returns:
        None (creates Excel file on disk)
    """
    # Check if file exists and is accessible
    if os.path.exists(output_filename):
        try:
            # Try to open the file in append mode to check if it's locked
            test_file = open(output_filename, 'r+b')
            test_file.close()
        except PermissionError:
            print(f"[ERROR] Error: The file '{output_filename}' is currently open in another application.")
            print("   Please close the file and try again.")
            sys.exit(1)
        except Exception as e:
            print(f"[ERROR] Error accessing file '{output_filename}': {e}")
            sys.exit(1)
    
    # Create Excel writer with error handling
    try:
        writer = pd.ExcelWriter(output_filename, engine="openpyxl", mode='w')
    except PermissionError:
        print(f"[ERROR] Error: Cannot write to '{output_filename}'. The file may be open in another application.")
        print("   Please close the file and try again.")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Error creating Excel writer: {e}")
        sys.exit(1)
    
    # Write dataframes to Excel
    for sheet_name, df in dataframes.items():
        # Excel sheet names max 31 chars
        safe_sheet_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
    
    # Save the workbook
    try:
        writer.close()
    except Exception as e:
        print(f"[ERROR] Error closing Excel writer: {e}")
        sys.exit(1)
    
    # Load workbook to add cover sheet and format tables
    print("Creating cover sheet and formatting tables...")
    try:
        workbook = load_workbook(output_filename)
        create_cover_sheet(workbook, output_filename, domains)
        create_documentation_sheet(workbook, dataframes)
        
        # Format all data sheets as tables and add hyperlinks
        for sheet_name, df in dataframes.items():
            safe_sheet_name = sheet_name[:31]
            if safe_sheet_name in workbook.sheetnames:
                worksheet = workbook[safe_sheet_name]
                format_sheet_as_table(worksheet, df)
                add_relationship_type_hyperlinks(worksheet, df)
                color_risk_column(worksheet, df)
        
        workbook.save(output_filename)
    except PermissionError:
        print(f"[ERROR] Error: Cannot save '{output_filename}'. The file may be open in another application.")
        print("   Please close the file and try again.")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Error saving workbook: {e}")
        sys.exit(1)
    
    print(f"[OK] Excel report generated: {output_filename}")
    
    # Generate risk calculation log if requested
    if enable_logging and risk_breakdowns:
        # Generate log filename based on output filename
        base_name = os.path.splitext(os.path.basename(output_filename))[0]
        log_filename = f"{base_name}_risk_calculation_log.txt"
        # Use the same directory as the output file, or current directory if output is just a filename
        if os.path.dirname(output_filename):
            log_filename = os.path.join(os.path.dirname(output_filename), log_filename)
        
        generate_risk_log(risk_breakdowns, output_filename, log_filename)


